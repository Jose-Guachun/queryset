from django.shortcuts import render, redirect

from crud.forms import ProductoForm
from crud.models import Producto
#importar para generar reportes en excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def list(request):
    productos = Producto.objects.all()
    return render(request,'productos/index.html',{'productos':productos})

class ReportePersonalizadoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        productos=Producto.objects.all()
        wb=Workbook()
        ws=wb.active
        #Estilo de titulo de cabecera
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B2'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
        ws['B2'].font = Font(name='Calibri', size=12, bold=True)
        ws['B2'] = 'REPORTE PERSONALIZADO EN EXCEL CON DJANGO'
        ws.merge_cells('B2:H2')

        #estilo de campos de cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].font = Font(name='Calibri', size=12, bold=True)

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].font = Font(name='Calibri', size=12, bold=True)

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].font = Font(name='Calibri', size=12, bold=True)

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].font = Font(name='Calibri', size=12, bold=True)

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].font = Font(name='Calibri', size=12, bold=True)

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].font = Font(name='Calibri', size=12, bold=True)

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].font = Font(name='Calibri', size=12, bold=True)

        #Nombre de Cabecera del excel
        ws['B3'] ='MARCA'
        ws['C3'] = 'MODELO'
        ws['D3'] = 'PRODUCTO'
        ws['E3'] = 'FECHA'
        ws['F3'] = 'CANTIDAD'
        ws['G3'] = 'PRECIO'
        ws['H3'] = 'TOTAL'

        cont=4
        ws.title = 'Hoja'
        for producto in productos:
            # Cambiar caracteristicas de las celdas
            ws.row_dimensions[1].height = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20

            # Pintamos los datos en el reporte
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=2).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=2).value = str(producto.marca)

            ws.cell(row=cont, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=3).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=3).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=3).value = str(producto.modelo)

            ws.cell(row=cont, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=4).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=4).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=4).value = producto.descripcion

            ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=5).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=5).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=5).value = producto.fecha

            ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=6).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=6).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=6).value = producto.cantidad

            ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=7).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=7).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=7).value = producto.precio

            ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=cont, column=8).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=cont, column=8).font = Font(name='Calibri', size=8)
            ws.cell(row=cont, column=8).value = producto.total

            cont+=1
        #establece el nombre de mi archivo
        nombre_archivo="ReporteExcel.xlsx"
        #definir el tipo de respuesta que se va a dar
        response=HttpResponse(content_type="application/ms-excel")
        contenido="attachment; filename= {0}".format(nombre_archivo)
        response["Content-Disposition"]=contenido
        wb.save(response)
        return response

def create(request):
    if request.method=='POST':
        form=ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            product = Producto()
            product.marca = form.cleaned_data['marca']
            product.modelo = form.cleaned_data['modelo']
            product.descripcion = form.cleaned_data['descripcion']
            product.fecha = form.cleaned_data['fecha']
            product.cantidad = form.cleaned_data['cantidad']
            product.precio = form.cleaned_data['precio']
            product.total =  product.cantidad*product.precio
            product.save()
            return redirect('crud:index')
    else:
        form=ProductoForm()

    return render(request,'productos/create.html',{'form':form})

def update(request,id):
    producto=Producto.objects.get(id=id)
    if request.method=='GET':
        form=ProductoForm(instance=producto)
    else:
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            product = Producto()
            product.marca = form.cleaned_data['marca']
            product.modelo = form.cleaned_data['modelo']
            product.descripcion = form.cleaned_data['descripcion']
            product.fecha = form.cleaned_data['fecha']
            product.cantidad = form.cleaned_data['cantidad']
            product.precio = form.cleaned_data['precio']
            product.total = product.cantidad * product.precio
            product.save()
            return redirect('crud:index')

    return render(request,'productos/create.html',{'form':form})

def delete(request,id):
    producto = Producto.objects.get(id=id)
    descripcion=producto.descripcion
    if request.method=='POST':
        producto.delete()
        return redirect('crud:index')
    return render(request,'productos/delete.html',{'descripcion':descripcion})

